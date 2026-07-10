import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse


async def read_rows(file: UploadFile) -> list[dict]:
    """Legge un file CSV o XLSX caricato e restituisce una lista di dict {intestazione: valore}."""
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        return []
    if filename.endswith(".xlsx"):
        return _read_xlsx(content)
    return _read_csv(content)


def _read_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = []
    for raw in reader:
        row = {}
        for k, v in raw.items():
            if not k:
                continue
            row[k.strip()] = v.strip() if isinstance(v, str) else v
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def _read_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "Supporto XLSX non disponibile sul server")
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for raw in rows_iter:
        if all(v is None for v in raw):
            continue
        row = {h: v for h, v in zip(headers, raw) if h}
        rows.append(row)
    return rows


def make_csv_template(filename: str, headers: list[str]) -> StreamingResponse:
    buf = io.StringIO()
    buf.write("﻿")  # BOM per apertura corretta in Excel
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _csv_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, Decimal):
        return format(v, "f")
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def make_csv_export(filename: str, headers: list[str], rows: list[dict]) -> StreamingResponse:
    """Esporta i dati correnti in un CSV con le stesse intestazioni del template di import,
    per consentire il flusso scarica -> modifica -> ricarica."""
    buf = io.StringIO()
    buf.write("﻿")  # BOM per apertura corretta in Excel
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_csv_value(row.get(h)) for h in headers])
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Conversioni di valore (gestiscono sia stringhe da CSV sia tipi nativi XLSX)
# ---------------------------------------------------------------------------

def str_or_none(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def decimal_or_none(v) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f"valore numerico non valido: {v!r}")


def bool_or_none(v, default: bool | None = None) -> bool | None:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in ("1", "true", "si", "sì", "x", "yes", "y", "vero"):
        return True
    if s in ("0", "false", "no", "n", "falso"):
        return False
    return default
